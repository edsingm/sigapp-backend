import openpyxl
wb = openpyxl.load_workbook('docs/viabilidade-modelo/viabilidade.xlsx', data_only=False)
ws = wb['Aux_Parcelas']

# Ver GY (col 209) para rows 21-30 (coortes de obra, block 1)
print("=== GY formulas rows 21-30 (block 1 - obra) ===")
for r in range(21, 31):
    gy = ws.cell(row=r, column=209).value  # GY
    gx = ws.cell(row=r, column=208).value  # GX
    gw = ws.cell(row=r, column=207).value  # GW
    gv = ws.cell(row=r, column=206).value  # GV
    gu = ws.cell(row=r, column=205).value  # GU
    gt = ws.cell(row=r, column=204).value  # GT
    gs = ws.cell(row=r, column=203).value  # GS
    g = ws.cell(row=r, column=7).value     # G (TOTAL)
    f = ws.cell(row=r, column=6).value     # F (PARCELA)
    e = ws.cell(row=r, column=5).value     # E (VENDAS)
    b = ws.cell(row=r, column=2).value     # B (mes)
    
    print(f"Row{r} (mes={b}): E(VENDAS)={e}, F(PARC)={f}, G(TOT)={g}")
    print(f"  GS={gs}, GT={gt}, GU={gu}, GV={gv}, GW={gw}, GX={gx}, GY={gy}")

# Ver a taxa - Premissas
ws2 = wb['Premissas']
print("\n=== Premissas - taxa juros ===")
for r in range(1, 60):
    for col in [3, 4, 5]:  # C, D, E = 2D, 3D, Lotes
        v = ws2.cell(row=r, column=col).value
        if v is not None:
            print(f"  Row{r} Col{col}: {v}")
